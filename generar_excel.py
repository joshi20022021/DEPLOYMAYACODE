import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

output_dir = r'C:\Users\edgar\Downloads\GEREN1\SOG1A_G14\Proyecto_Fase2\Entregables_Excel'
os.makedirs(output_dir, exist_ok=True)

header_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

def style_headers(ws, n_cols):
    for col in range(1, n_cols + 1):
        ws.cell(1, col).fill = header_fill
        ws.cell(1, col).font = header_font
        ws.cell(1, col).alignment = Alignment(horizontal='center')

def set_col_widths(ws, width, n_cols):
    for col in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = width

# ---- PRODUCTOS ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Productos'
headers = ['Codigo', 'Nombre', 'Categoria', 'Precio Venta (Q)', 'Costo (Q)', 'Tipo', 'Proveedor Principal']
ws.append(headers)
style_headers(ws, len(headers))

products_data = [
    ('DELL-INS15', 'Dell Inspiron 15 3000', 'Laptops', 4500, 3800, 'Bien', 'Dell Technologies'),
    ('DELL-LAT54', 'Dell Latitude 5420', 'Laptops', 7200, 6000, 'Bien', 'Dell Technologies'),
    ('DELL-XPS13', 'Dell XPS 13', 'Laptops', 10500, 8800, 'Bien', 'Dell Technologies'),
    ('HP-PAV15', 'HP Pavilion 15', 'Laptops', 4200, 3500, 'Bien', 'HP Inc.'),
    ('HP-ELI840', 'HP EliteBook 840 G8', 'Laptops', 8900, 7400, 'Bien', 'HP Inc.'),
    ('HP-PRO450', 'HP ProBook 450 G8', 'Laptops', 5800, 4800, 'Bien', 'HP Inc.'),
    ('DELL-U2422', 'Dell UltraSharp U2422H', 'Monitores', 2800, 2200, 'Bien', 'Dell Technologies'),
    ('DELL-P2422', 'Dell P2422H Monitor', 'Monitores', 1900, 1500, 'Bien', 'Dell Technologies'),
    ('SAM-27CF', 'Samsung 27 Curved FHD', 'Monitores', 2200, 1750, 'Bien', 'Samsung Electronics'),
    ('SAM-ODG532', 'Samsung Odyssey G5 32', 'Monitores', 4500, 3600, 'Bien', 'Samsung Electronics'),
    ('LOG-MXM3', 'Logitech MX Master 3', 'Perifericos', 650, 480, 'Bien', 'Logitech Guatemala'),
    ('LOG-MK470', 'Logitech MK470 Combo', 'Perifericos', 450, 320, 'Bien', 'Logitech Guatemala'),
    ('LOG-G502', 'Logitech G502 Hero Mouse', 'Perifericos', 580, 420, 'Bien', 'Logitech Guatemala'),
    ('LOG-K380', 'Logitech K380 Teclado', 'Perifericos', 320, 230, 'Bien', 'Logitech Guatemala'),
    ('HP-WC720', 'HP Webcam HD 720p', 'Perifericos', 280, 190, 'Bien', 'HP Inc.'),
    ('KST-SSD240', 'Kingston SSD A400 240GB', 'Almacenamiento', 320, 240, 'Bien', 'Kingston Technology'),
    ('KST-SSD480', 'Kingston SSD A400 480GB', 'Almacenamiento', 520, 400, 'Bien', 'Kingston Technology'),
    ('KST-NV21TB', 'Kingston NV2 1TB NVMe', 'Almacenamiento', 780, 600, 'Bien', 'Kingston Technology'),
    ('KST-RAM16', 'Kingston RAM DDR4 16GB', 'Almacenamiento', 420, 320, 'Bien', 'Kingston Technology'),
    ('SAM-EVO1TB', 'Samsung 870 EVO 1TB SSD', 'Almacenamiento', 950, 750, 'Bien', 'Samsung Electronics'),
    ('SAM-980P512', 'Samsung 980 Pro 512GB', 'Almacenamiento', 820, 650, 'Bien', 'Samsung Electronics'),
    ('TPL-AX50', 'TP-Link Archer AX50', 'Redes', 850, 650, 'Bien', 'TP-Link Guatemala'),
    ('TPL-SG108', 'TP-Link TL-SG108 Switch', 'Redes', 320, 240, 'Bien', 'TP-Link Guatemala'),
    ('TPL-RE500X', 'TP-Link RE500X Extensor', 'Redes', 450, 340, 'Bien', 'TP-Link Guatemala'),
    ('TPL-DECOX20', 'TP-Link Deco X20 Mesh', 'Redes', 1200, 950, 'Bien', 'TP-Link Guatemala'),
    ('DELL-MOCH15', 'Mochila Dell Pro 15', 'Accesorios', 380, 280, 'Bien', 'Dell Technologies'),
    ('DELL-WD19S', 'Docking Station Dell WD19S', 'Accesorios', 2200, 1750, 'Bien', 'Dell Technologies'),
    ('LOG-C920', 'Logitech C920 Webcam HD', 'Accesorios', 580, 430, 'Bien', 'Logitech Guatemala'),
    ('ACC-HUBC7', 'Hub USB-C 7 en 1', 'Accesorios', 320, 230, 'Bien', 'Dell Technologies'),
    ('MC-PADXL', 'Pad Mouse XL MayaCode', 'Accesorios', 120, 80, 'Bien', 'Logitech Guatemala'),
    ('ACC-SOPL', 'Soporte Laptop Aluminio', 'Accesorios', 250, 180, 'Bien', 'Dell Technologies'),
]
for row in products_data:
    ws.append(row)
set_col_widths(ws, 24, len(headers))
wb.save(os.path.join(output_dir, 'productos_MayaCode.xlsx'))
print('productos_MayaCode.xlsx generado -', len(products_data), 'productos')

# ---- PROVEEDORES ----
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Proveedores'
headers2 = ['Nombre', 'Email', 'Telefono', 'Direccion', 'Ciudad', 'Pais', 'Categoria Productos']
ws2.append(headers2)
style_headers(ws2, len(headers2))

proveedores_data = [
    ('Dell Technologies', 'ventas@dell.com', '+1-800-999-3355', '1 Dell Way', 'Round Rock', 'Estados Unidos', 'Laptops, Monitores, Accesorios'),
    ('HP Inc.', 'ventas@hp.com', '+1-800-474-6836', '1501 Page Mill Rd', 'Palo Alto', 'Estados Unidos', 'Laptops, Perifericos'),
    ('Logitech Guatemala', 'ventas@logitech.gt', '+502 2200-0100', '6a Avenida 10-20', 'Guatemala', 'Guatemala', 'Perifericos, Accesorios'),
    ('Kingston Technology', 'sales@kingston.com', '+1-714-435-2600', '17600 Newhope St', 'Fountain Valley', 'Estados Unidos', 'Almacenamiento, RAM'),
    ('TP-Link Guatemala', 'ventas@tplink.gt', '+502 2222-3344', '12 Calle 4-55', 'Guatemala', 'Guatemala', 'Redes'),
    ('Samsung Electronics', 'b2b@samsung.com', '+1-800-726-7864', '85 Challenger Rd', 'Ridgefield Park', 'Estados Unidos', 'Monitores, Almacenamiento'),
]
for row in proveedores_data:
    ws2.append(row)
set_col_widths(ws2, 28, len(headers2))
wb2.save(os.path.join(output_dir, 'proveedores_MayaCode.xlsx'))
print('proveedores_MayaCode.xlsx generado -', len(proveedores_data), 'proveedores')

# ---- CLIENTES ----
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = 'Clientes'
headers3 = ['Nombre', 'Email', 'Telefono', 'Direccion', 'Ciudad', 'Tipo', 'Segmento']
ws3.append(headers3)
style_headers(ws3, len(headers3))

clientes_data = [
    ('Tecnologias Avanzadas S.A.', 'compras@tecnoav.gt', '+502 2200-1100', '5a Avenida 12-34', 'Guatemala', 'Empresa', 'Corporativo'),
    ('Universidad del Valle de Guatemala', 'compras@uvg.edu.gt', '+502 2369-0791', '18 Calle 11-95', 'Guatemala', 'Empresa', 'Educacion'),
    ('Banco Industrial', 'tecnologia@bi.com.gt', '+502 2420-3000', '7a Avenida 5-10', 'Guatemala', 'Empresa', 'Financiero'),
    ('Grupo Progreso', 'it@progreso.com.gt', '+502 2323-7000', '1a Calle 7-66', 'Mixco', 'Empresa', 'Corporativo'),
    ('Clinicas y Hospitales La Familiar', 'sistemas@lafamiliar.gt', '+502 2500-4000', '15 Avenida 12-20', 'Guatemala', 'Empresa', 'Salud'),
    ('Juan Carlos Mendez', 'jcmendez@gmail.com', '+502 5555-1234', '10a Calle 8-50', 'Quetzaltenango', 'Persona', 'Particular'),
    ('Maria Elena Lopez', 'melopez@outlook.com', '+502 5544-2233', '3a Avenida 5-30', 'Antigua Guatemala', 'Persona', 'Particular'),
    ('Roberto Giron Castillo', 'rgiron@yahoo.com', '+502 5678-9012', '6a Calle 2-15', 'Escuintla', 'Persona', 'Particular'),
    ('Ana Lucia Ramirez', 'alramirez@gmail.com', '+502 5432-1100', '2a Avenida 9-40', 'Coban', 'Persona', 'Particular'),
    ('Colegios Americanos de Guatemala', 'admin@colegiosamericanos.edu.gt', '+502 2383-4444', '20 Calle 5-60', 'Guatemala', 'Empresa', 'Educacion'),
    ('Distribuidora El Exito', 'gerencia@elexito.gt', '+502 2277-8800', '9a Calle 15-22', 'Guatemala', 'Empresa', 'Comercio'),
    ('Carlos Herrera Fuentes', 'cherrera@hotmail.com', '+502 5901-3344', '11a Avenida 3-10', 'Chiquimula', 'Persona', 'Particular'),
    ('Ingenieria y Construccion SA', 'it@icsagt.com', '+502 2290-5500', '14 Avenida 3-88', 'Villa Nueva', 'Empresa', 'Construccion'),
    ('Sandra Patricia Ajanel', 'spajanel@gmail.com', '+502 5812-6677', '8a Calle 7-25', 'Huehuetenango', 'Persona', 'Particular'),
    ('Municipalidad de Guatemala', 'informatica@muniguate.com', '+502 2285-0000', '21 Calle 6-77', 'Guatemala', 'Empresa', 'Gobierno'),
]
for row in clientes_data:
    ws3.append(row)
set_col_widths(ws3, 30, len(headers3))
wb3.save(os.path.join(output_dir, 'clientes_MayaCode.xlsx'))
print('clientes_MayaCode.xlsx generado -', len(clientes_data), 'clientes')
print('Listo! Archivos en:', output_dir)
